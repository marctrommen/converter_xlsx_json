#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl

import logging
logger = logging.getLogger("myapp.ExcelReader")


# -----------------------------------------------------------------------------
class ExcelReader:
    """This class is responsible for reading Excel files (.xlsx) and extracting data
    from worksheets that start with 'SST_'."""


    # -----------------------------------------------------------------------------
    def __init__(self) -> None:
        """Initialize the ExcelReader."""
        
        logger.debug("ExcelReader initialized")
        self.error_code = 0


    # -----------------------------------------------------------------------------
    def read_excel_file(self, 
                        excel_file_path: str, 
                        worksheet_prefix: str, 
                        excel_data: dict) -> int:
        """Read from an Excel workbook located by 'excel_file_path'.
         Iterates through all worksheets to filter them by name, starting 
         with 'worksheet_prefix'. Reads all data fields and stores them in 'excel_data'.
         Returns 0 on success, otherwise returns an error code."""
        
        logger.debug("Reading Excel file: %s", excel_file_path)

        self.error_code = 0

        try:
            workbook = openpyxl.load_workbook(filename=excel_file_path, read_only=True)
        except Exception as e:
            logger.error("Error reading Excel file '%s': %s", excel_file_path, e)
            self.error_code = 11
        
        if self.error_code == 0:
            for sheet_name in workbook.sheetnames:
                if sheet_name.startswith(worksheet_prefix):
                    logger.debug("Processing worksheet: %s", sheet_name)

                    worksheet = workbook[sheet_name]
                    worksheet_data = {}
                    worksheet_id = f"{excel_file_path}:{sheet_name}"

                    self._read_data_from_worksheet(worksheet, worksheet_data)
                    excel_data[worksheet_id] = worksheet_data

        return self.error_code


    # -----------------------------------------------------------------------------
    def _read_data_from_worksheet(self, 
                                  worksheet:openpyxl.worksheet.worksheet.Worksheet,
                                  sheet_data:dict) -> None:
        """This method iterates through all rows in the worksheet and appends the 
        data to the provided dictionary.
        It assumes that the first row contains headers and subsequent rows contain data.
        It assumes that the first column is the identifier, in the second column is 
        the value for each row.
        It iterates If the first column is empty, it will stop iterating through the skip that row."""

        logger.debug("Reading data from worksheet")

        for row in range(2, worksheet.max_row + 1):
            name = worksheet.cell(row=row, column=1).value
            value = worksheet.cell(row=row, column=2).value

            if name is None:
                logger.debug("Empty name found at row %d, stopping iteration", row)
                break
            name = str(name).strip()

            data_key = ""

            if value is None:
                logger.debug("Empty value found at row %d", row)
                value = ""
            else:
                value = str(value).strip()

            if name == "Schnittstellenname":
                data_key = "sst_name"
            elif name == "Versionierung":
                data_key = "sst_version"
            elif name == "Gültigkeitszeitraum von":
                data_key = "sst_valid_from"
            elif name == "Gültigkeitszeitraum bis":
                data_key = "sst_valid_to"
            elif name == "Schnittstelle aktiv":
                data_key = "sst_is_active"
            elif name == "Verantwortliche Stelle/Kontakt":
                data_key = "sst_responsible"
            elif name == "Authentifizierungsmethoden":
                data_key = "sst_auth_method"
            elif name == "Autorisierungsmechanismen":
                data_key = "sst_authorization"
            elif name == "Verschlüsselung":
                data_key = "sst_crypto"
            elif name == "Datenformate":
                data_key = "sst_format"
            elif name == "Trigger":
                data_key = "sst_trigger"
            elif name == "Austauschprotokoll":
                data_key = "sst_protocol"
            elif name == "Transportmechanismus":
                data_key = "sst_transport"
            elif name == "Schnittstelle_1_System":
                data_key = "partner_1_system"
            elif name == "Schnittstelle_1_Verantwortlicher":
                data_key = "partner_1_responsible"
            elif name == "Schnittstelle_2_System":
                data_key = "partner_2_system"
            elif name == "Schnittstelle_2_Verantwortlicher":
                data_key = "partner_2_responsible"
            else:
                logger.debug("Unknown name '%s', skipping", name)
                continue
            
            sheet_data[data_key] = value
            logger.debug("Added data: %s = %s", data_key, value)


